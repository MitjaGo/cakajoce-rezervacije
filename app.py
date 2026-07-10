"""
Streamlit aplikacija: Filtriranje rezervacij s statusom "Na čakanju"
======================================================================
Naloži 1-6 XLS izvoznih datotek (PMS sistem, HTML-tabela s pripono .xls),
filtrira vrstice s statusom "Na čakanju", kjer je od stolpca
"Datum nastanka" do referenčnega datuma preteklo N ali več dni (privzeto 4).
Referenčni datum NI sistemski "danes" (ker je lahko ura strežnika napačna),
ampak se izračuna samodejno iz podatkov - kot najnovejši datum, ki se
pojavi v stolpcu "Datum nastanka" med vsemi naloženimi datotekami. Aplikacija
združi rezultate vseh datotek v en Excel dokument in omogoči prenos na
računalnik.

Zagon:
    pip install -r requirements.txt
    streamlit run app.py
"""

import re
from datetime import date
from io import BytesIO
from urllib.parse import quote

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Rezervacije - Na čakanju", layout="wide")

LOGO_URL = "https://www.adria-ankaran.si//app/uploads/2025/10/logo-Adria.jpg"

header_left, header_right = st.columns([4, 1])
with header_left:
    st.title("📋 Rezervacije s statusom \"Na čakanju\"")
with header_right:
    st.image(LOGO_URL, width=120)

st.markdown(
    """
Naloži od **1 do 6** XLS datotek (izvoz iz sistema PHOBS / Rezervacije na čakanju, označi pred prenosom v excel samo rezervacije s statusom na čakanju obarvane z oranžno ali rumeno barvo ). 

- Aplikacija bo prebrala podatke iz vsake datoteke,
- samodejno ugotovila referenčni datum ("danes") kot **najnovejši datum v
  stolpcu Datum nastanka** med vsemi naloženimi podatki (ne zanaša se na
  sistemsko uro strežnika),
- prikazala **vse vrstice s statusom Na čakanju**, barvno pa označila
  nujnost glede na to, koliko dni je preteklo od stolpca **Datum nastanka**
  do referenčnega datuma,
- prikazala stolpce: **Številka PH, HIS, Objekt, Datum ponudbe, Prihod,
  Lastnik rezervacije, Status**,
- združila rezultate vseh naloženih datotek v en Excel dokument, ki ga
  prenesete na svoj računalnik kot xls datoteko, kopirate v odložišče za pošiljanje v mailu (tako da odprete nov e-mail in pritisnite ctrl+V ali sprintate kot A4 format.
"""
)

# ---------------------------------------------------------------------------
# Nastavitve filtra
# ---------------------------------------------------------------------------
URGENT_DAYS = 3  # rezervacije s prihodom 1, 2 ali 3 dni po nastanku - vedno prikazane
LONG_LEAD_DAYS = 10  # rezervacije s prihodom >10 dni po nastanku - za spremljati (modra)

narrow_col, _spacer = st.columns([1, 3])
with narrow_col:
    min_days = st.number_input(
        "(Koliko dni je minimalno rezervacija na čakanju, privzeto 4 ali več)", min_value=0, value=4, step=1
    )

st.caption(
    "Prikazane so VSE vrstice s statusom 'Na čakanju'. Barva pove nujnost: "
    f"🔴 rdeča = prihod je {URGENT_DAYS} dni ali manj od nastanka (gost mora "
    "plačati vnaprej, nujno preveriti); 🔵 svetlo modra = prihod je več kot "
    f"{LONG_LEAD_DAYS} dni od nastanka (ni nujno, a naj se preveri plačilo "
    "pred prihodom); brez barve = na čakanju ≥ zgornji prag dni ali vmesno "
    "obdobje."
)

uploaded_files = st.file_uploader(
    "Naloži XLS datoteke (1-6 datotek)",
    type=["xls", "xlsx"],
    accept_multiple_files=True,
)

REQUIRED_COLS = [
    "Status",
    "PMS koda",
    "Code",
    "Objekt",
    "Datum nastanka",
    "Prihod",
    "Lastnik rezervacije",
]


# ---------------------------------------------------------------------------
# Pomožne funkcije
# ---------------------------------------------------------------------------
def _pick_best_table(content: bytes):
    """PMS izvoz ima v prvi vrstici naslovni <td colspan=...> (npr. 'Premium
    mobile homes'), zato mora biti pravi header v vrstici 1, ne 0. Preizkusi
    header=0 in header=1 ter obdrži različico, ki vsebuje pričakovane
    stolpce (ali ima v vsakem primeru največ prepoznanih stolpcev)."""
    best_df = None
    best_score = -1
    for header_row in (1, 0, None):
        try:
            tables = pd.read_html(BytesIO(content), header=header_row)
        except Exception:
            continue
        if not tables:
            continue
        candidate = max(tables, key=lambda t: t.shape[1])
        candidate.columns = [str(c).strip() for c in candidate.columns]
        score = sum(1 for c in REQUIRED_COLS if c in candidate.columns)
        if score > best_score:
            best_score = score
            best_df = candidate
        if score == len(REQUIRED_COLS):
            break
    return best_df


def parse_file(file) -> "pd.DataFrame | None":
    """Datoteke so v resnici HTML tabela shranjena s pripono .xls."""
    content = file.read()
    df = _pick_best_table(content)
    if df is None:
        # poskusi kot pravi binarni/OOXML Excel
        try:
            file.seek(0)
            return pd.read_excel(file)
        except Exception as e2:
            st.error(f"Napaka pri branju datoteke {file.name}: {e2}")
            return None
    return df


def extract_number(val):
    if pd.isna(val):
        return None
    m = re.search(r"\d+", str(val))
    return m.group() if m else None


def parse_date(val):
    if pd.isna(val):
        return None
    s = str(val).strip()
    if not s:
        return None
    try:
        return pd.to_datetime(s).date()
    except Exception:
        return None


def filter_dataframe(df: pd.DataFrame, file_name: str, filter_date, min_days, urgent_days, long_lead_days) -> "pd.DataFrame | None":
    """Filtrira že prebran DataFrame (glej parse_file) glede na status
    'Na čakanju' in datumske pogoje, relativno na podan filter_date."""
    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        st.error(f"Datoteka **{file_name}** nima pričakovanih stolpcev: {missing}")
        return None

    work = df[REQUIRED_COLS].copy()

    # razpar-anje datumov
    work["_datum_nastanka_parsed"] = work["Datum nastanka"].apply(parse_date)
    work["_prihod_parsed"] = work["Prihod"].apply(parse_date)
    work = work.dropna(subset=["_datum_nastanka_parsed"])

    # filter statusa - "Na čakanju" (case-insensitive, robustno na HTML/presledke)
    work["_status_clean"] = work["Status"].astype(str).str.strip()
    work = work[work["_status_clean"].str.contains("čakanju", case=False, na=False)]

    if work.empty:
        return work

    # dni od nastanka do referenčnega datuma (dolgo čakanje)
    work["Dni od nastanka"] = work["_datum_nastanka_parsed"].apply(
        lambda d: (filter_date - d).days
    )

    # dni med nastankom rezervacije in prihodom gosta (prihod kmalu = nujno,
    # ker gost mora plačati vnaprej, rok za urejanje je kratek)
    work["Dni do prihoda (od nastanka)"] = work.apply(
        lambda r: (r["_prihod_parsed"] - r["_datum_nastanka_parsed"]).days
        if pd.notna(r["_prihod_parsed"])
        else None,
        axis=1,
    )

    # Prikažemo VSE vrstice s statusom 'Na čakanju', ne glede na dneve.
    # Barvna kategorizacija (rdeča/modra/brez) spodaj ostane kot informativna
    # oznaka nujnosti, a ne omejuje več, katere vrstice se sploh prikažejo.
    if work.empty:
        return work

    def _razlog(row):
        r = []
        if row["Dni od nastanka"] >= min_days:
            r.append(f"Na čakanju enako ali več kot ({min_days} dni)")
        d = row["Dni do prihoda (od nastanka)"]
        if d is not None and d <= urgent_days:
            r.append(f"Prihod kmalu čez (1,2 {urgent_days} dni)")
        if d is not None and d > long_lead_days:
            r.append(f"Pridejo čez več kot {long_lead_days} dni (preveri plačilo)")
        if not r:
            r.append("V vmesnem obdobju (spremljaj)")
        return " + ".join(r)

    work["Razlog"] = work.apply(_razlog, axis=1)

    work["PMS koda"] = work["PMS koda"].apply(extract_number)
    work["Code"] = work["Code"].astype(str).str.strip()
    work["Objekt"] = work["Objekt"].astype(str).str.strip()
    work["Datum nastanka"] = work["_datum_nastanka_parsed"].astype(str)
    work["Prihod"] = work["Prihod"].astype(str).str.strip()
    work["Lastnik rezervacije"] = work["Lastnik rezervacije"].astype(str).str.strip()
    work["Vir datoteke"] = file_name

    final_cols = [
        "Code",
        "PMS koda",
        "Objekt",
        "Datum nastanka",
        "Prihod",
        "Lastnik rezervacije",
        "Status",
        "Dni od nastanka",
        "Dni do prihoda (od nastanka)",
        "Razlog",
    ]
    result = work[final_cols].reset_index(drop=True)

    # preimenovanje za prikaz/izvoz (interno branje iz XLS ostane nespremenjeno)
    result = result.rename(
        columns={
            "Code": "Številka PH",
            "PMS koda": "HIS",
            "Datum nastanka": "Datum nastanka (ponudbe)",
            "Dni od nastanka": "Število preteklih dni (od nastanka)",
        }
    )
    return result


def _row_color(razlog: str) -> "str | None":
    """Vrne barvo vrstice glede na vsebino stolpca Razlog:
    - 'red'  - nujno (prihod kmalu po nastanku, ≤ URGENT_DAYS dni)
    - 'blue' - za spremljati (prihod precej oddaljen, > LONG_LEAD_DAYS dni)
    - None   - brez posebne barve (samo dolgo čakanje, brez drugih pogojev)
    """
    s = str(razlog)
    if "Prihod kmalu" in s:
        return "red"
    if "Pridejo čez" in s:
        return "blue"
    return None


_COLOR_HEX = {"red": "#ffcccc", "blue": "#cce5ff"}


def _table_html_parts(df: pd.DataFrame, color_series: pd.Series):
    """Vrne (header_html, rows_html) - skupna gradnja za tisk in kopiranje."""
    header_html = "".join(f"<th>{c}</th>" for c in df.columns)
    rows_html = ""
    for idx, row in df.iterrows():
        color = color_series.loc[idx]
        row_style = f' style="background-color:{_COLOR_HEX[color]};"' if color else ""
        cells = "".join(f"<td>{'' if pd.isna(v) else v}</td>" for v in row)
        rows_html += f"<tr{row_style}>{cells}</tr>"
    return header_html, rows_html


def build_print_html(df: pd.DataFrame, color_series: pd.Series, filter_date) -> str:
    """Zgradi samostojen HTML dokument s tabelo, oblikovan za tiskanje na A4,
    z gumbom, ki sproži tiskanje (window.print())."""
    header_html, rows_html = _table_html_parts(df, color_series)

    return f"""
    <html>
    <head>
    <meta charset="utf-8">
    <style>
        @page {{ size: A4 landscape; margin: 12mm; }}
        body {{ font-family: Arial, Helvetica, sans-serif; }}
        h2 {{ margin: 0 0 4px 0; }}
        p.meta {{ margin: 0 0 12px 0; color: #555; font-size: 12px; }}
        table {{ border-collapse: collapse; width: 100%; }}
        th, td {{ border: 1px solid #999; padding: 4px 6px; font-size: 10px; text-align: left; }}
        th {{ background-color: #f0f0f0; }}
        #printBtn {{
            padding: 8px 18px; font-size: 14px; cursor: pointer;
            background-color: #d63333; color: white; border: none; border-radius: 4px;
        }}
        @media print {{
            #printBtn {{ display: none; }}
        }}
    </style>
    </head>
    <body>
        <button id="printBtn" onclick="window.print()">🖨️ Natisni na A4</button>
        <h2>Rezervacije - Na čakanju</h2>
        <p class="meta">Referenčni datum: {filter_date} · Skupno vrstic: {len(df)}</p>
        <table>
            <thead><tr>{header_html}</tr></thead>
            <tbody>{rows_html}</tbody>
        </table>
    </body>
    </html>
    """


def build_table_html_for_clipboard(df: pd.DataFrame, color_series: pd.Series, filter_date) -> str:
    """Zgradi HTML tabelo (brez gumbov/strani), primerno za kopiranje v
    odložišče in lepljenje neposredno v telo e-maila (npr. Outlook), kjer se
    prikaže kot prava, oblikovana tabela - enako kot pri tisku."""
    header_html, rows_html = _table_html_parts(df, color_series)
    return (
        f'<div style="font-family:Arial,Helvetica,sans-serif;">'
        f'<h3 style="margin:0 0 4px 0;">Rezervacije - Na čakanju</h3>'
        f'<p style="margin:0 0 10px 0;color:#555;font-size:12px;">'
        f'Referenčni datum: {filter_date} &middot; Skupno vrstic: {len(df)}</p>'
        f'<table style="border-collapse:collapse;width:100%;">'
        f'<thead><tr>{header_html.replace("<th>", "<th style=\'border:1px solid #999;padding:4px 6px;background:#f0f0f0;font-size:11px;text-align:left;\'>")}</tr></thead>'
        f'<tbody>{rows_html.replace("<td>", "<td style=\'border:1px solid #999;padding:4px 6px;font-size:11px;\'>")}</tbody>'
        f'</table></div>'
    )


def build_table_text_for_clipboard(df: pd.DataFrame) -> str:
    """Navadno-tekstovna (tab-ločena) različica tabele - kot rezervni format
    za odložišče (npr. za lepljenje v Excel)."""
    lines = ["\t".join(str(c) for c in df.columns)]
    for _, row in df.iterrows():
        lines.append("\t".join("" if pd.isna(v) else str(v) for v in row))
    return "\n".join(lines)


def build_mailto_link(df: pd.DataFrame, filter_date, recipient: str = "") -> str:
    """Zgradi mailto: povezavo. Ker mailto ne podpira HTML telesa, doda
    napotek za lepljenje predhodno kopirane tabele (gumb 'Kopiraj tabelo'),
    pod njim pa še preprost tekstovni povzetek kot rezervo."""
    subject = f"Rezervacije na čakanju - {filter_date}"
    lines = [
        "Tukaj prilepi tabelo (Ctrl+V) - najprej klikni gumb 'Kopiraj tabelo':",
        "",
        "",
        "---",
        f"Rezervacije s statusom 'Na čakanju' na dan {filter_date} "
        f"(skupno {len(df)} vrstic) - tekstovni povzetek:",
        "",
    ]
    for _, row in df.iterrows():
        lines.append(
            f"{row.get('Številka PH', '')} | {row.get('HIS', '')} | {row.get('Objekt', '')} | "
            f"{row.get('Lastnik rezervacije', '')} | {row.get('Razlog', '')}"
        )
    body = "\n".join(lines)
    return f"mailto:{recipient}?subject={quote(subject)}&body={quote(body)}"


# ---------------------------------------------------------------------------
# Glavna logika
# ---------------------------------------------------------------------------
if uploaded_files:
    if len(uploaded_files) > 6:
        st.warning("Naložiš lahko največ 6 datotek. Upoštevanih bo prvih 6.")
        uploaded_files = uploaded_files[:6]

    # 1. korak: preberi vse datoteke enkrat (izognemo se dvojnemu branju)
    parsed_files = []  # [(ime_datoteke, df), ...]
    for f in uploaded_files:
        df = parse_file(f)
        if df is None:
            continue
        missing = [c for c in REQUIRED_COLS if c not in df.columns]
        if missing:
            st.error(f"Datoteka **{f.name}** nima pričakovanih stolpcev: {missing}")
            continue
        parsed_files.append((f.name, df))

    # 2. korak: izračunaj referenčni datum ("danes") iz podatkov samih -
    # najnovejši datum v stolpcu 'Datum nastanka' med vsemi naloženimi
    # datotekami. To se NE zanaša na (morda napačno) sistemsko uro strežnika.
    all_creation_dates = []
    for _, df in parsed_files:
        if "Datum nastanka" in df.columns:
            parsed_dates = df["Datum nastanka"].apply(parse_date).dropna()
            all_creation_dates.extend(parsed_dates.tolist())

    if all_creation_dates:
        filter_date = max(all_creation_dates)
    else:
        filter_date = date.today()  # rezerva, če v podatkih ni najdenega datuma

    st.info(
        f"📅 Referenčni datum filtracije (najnovejši 'Datum nastanka' v "
        f"naloženih podatkih): **{filter_date}**"
    )

    # 3. korak: filtriraj vsako datoteko glede na izračunani filter_date
    all_results = []
    for file_name, df in parsed_files:
        res = filter_dataframe(df, file_name, filter_date, min_days, URGENT_DAYS, LONG_LEAD_DAYS)
        if res is not None and not res.empty:
            all_results.append(res)
            st.caption(f"✅ {file_name}: najdenih {len(res)} vrstic")
        elif res is not None:
            st.caption(f"⚪ {file_name}: ni vrstic, ki ustrezajo pogojem")

    if all_results:
        combined = pd.concat(all_results, ignore_index=True)
        combined.index = combined.index + 1  # zaporedna številka naj se začne pri 1
        st.success(f"Skupno najdenih {len(combined)} vrstic, ki ustrezajo pogojem.")

        color_series = combined["Razlog"].astype(str).apply(_row_color)

        def _highlight_row(row):
            color = color_series.loc[row.name]
            css = f"background-color: {_COLOR_HEX[color]}" if color else ""
            return [css for _ in row]

        st.caption(
            "🔴 Rdeče = prihod je 1-3 dni od nastanka rezervacije - nujno preveriti. "
            f"🔵 Svetlo modro = prihod je več kot {LONG_LEAD_DAYS} dni od nastanka - "
            "ni nujno, a naj se preveri plačilo pred prihodom."
        )
        st.dataframe(
            combined.style.apply(_highlight_row, axis=1),
            use_container_width=True,
        )

        # Excel za prenos (z rdečo/modro osvetlitvijo vrstic glede na kategorijo)
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            combined.to_excel(writer, index=False, sheet_name="Na čakanju")
            worksheet = writer.sheets["Na čakanju"]
            fills = {
                "red": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
                "blue": PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
            }
            n_cols = combined.shape[1]
            for excel_row, color in enumerate(color_series, start=2):  # vrstica 1 = header
                if color:
                    for col in range(1, n_cols + 1):
                        worksheet.cell(row=excel_row, column=col).fill = fills[color]
        output.seek(0)

        btn_col1, btn_col2, btn_col3 = st.columns(3)
        with btn_col1:
            st.download_button(
                label="⬇️ Prenesi kot Excel (.xlsx)",
                data=output,
                file_name=f"na_cakanju_{filter_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with btn_col2:
            print_html = build_print_html(combined, color_series, filter_date)
            components.html(
                f"""
                <div style="display:flex; justify-content:center;">
                    <button onclick="printTable()" style="
                        width:100%; padding:8px 0; font-size:14px; cursor:pointer;
                        background-color:#31333F; color:white; border:none;
                        border-radius:6px;">🖨️ Natisni (A4)</button>
                </div>
                <script>
                function printTable() {{
                    var w = window.open('', '_blank');
                    w.document.write({print_html!r});
                    w.document.close();
                    w.focus();
                    setTimeout(function() {{ w.print(); }}, 300);
                }}
                </script>
                """,
                height=45,
            )
        with btn_col3:
            copy_html = build_table_html_for_clipboard(combined, color_series, filter_date)
            copy_text = build_table_text_for_clipboard(combined)
            components.html(
                f"""
                <div style="display:flex; justify-content:center;">
                    <button id="copyBtn" onclick="copyTable()" style="
                        width:100%; padding:8px 0; font-size:14px; cursor:pointer;
                        background-color:#197935; color:white; border:none;
                        border-radius:6px;">📋 Kopiraj tabelo v odložišče</button>
                </div>
                <script>
                async function copyTable() {{
                    var btn = document.getElementById('copyBtn');
                    var htmlContent = {copy_html!r};
                    var textContent = {copy_text!r};
                    try {{
                        var item = new ClipboardItem({{
                            'text/html': new Blob([htmlContent], {{type: 'text/html'}}),
                            'text/plain': new Blob([textContent], {{type: 'text/plain'}})
                        }});
                        await navigator.clipboard.write([item]);
                        btn.innerText = '✅ Kopirano!';
                    }} catch (err) {{
                        btn.innerText = '⚠️ Kopiranje ni uspelo';
                        console.error(err);
                    }}
                    setTimeout(function() {{ btn.innerText = '📋 Kopiraj tabelo'; }}, 2500);
                }}
                </script>
                """,
                height=45,
            )

        st.caption(
            "💡 Za lepo oblikovano tabelo v e-mailu: klikni **'📋 Kopiraj "
            "tabelo v odložišče'**, nato odpri nov e-mail (npr. v Outlooku) "
            "in v telo prilepi (Ctrl+V) - tabela se prilepi enako oblikovana "
            "kot pri tisku."
        )
    else:
        st.info("Ni najdenih vrstic, ki bi ustrezale filtru v nobeni naloženi datoteki.")
else:
    st.info("Prosim, naloži vsaj eno XLS datoteko (do največ 6).")
